from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook, load_workbook

import pandas as pd

from file_processor.cleaning import clean_numeric_columns, clean_text_columns, validate_required_columns
from file_processor.excel_builder import write_excel_report
from file_processor.pipelines.online_retail import OnlineRetailReportConfig, build_online_retail_report


def test_cleaning_helpers_standardize_common_columns() -> None:
    frame = pd.DataFrame({"Name": [" Alpha ", None], "Amount": ["10.5", "bad"]})

    validate_required_columns(frame, ["Name", "Amount"])
    cleaned = clean_text_columns(frame, ["Name"])
    cleaned = clean_numeric_columns(cleaned, ["Amount"])

    assert cleaned.loc[0, "Name"] == "Alpha"
    assert cleaned.loc[1, "Name"] == ""
    assert cleaned.loc[0, "Amount"] == 10.5
    assert pd.isna(cleaned.loc[1, "Amount"])


def test_excel_builder_writes_formatted_workbook(tmp_path) -> None:
    output_path = tmp_path / "report.xlsx"

    write_excel_report({"Summary": pd.DataFrame([{"Metric": "Rows", "Value": 2}])}, output_path)

    workbook = load_workbook(output_path, read_only=False)
    worksheet = workbook["Summary"]
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref is not None
    assert worksheet["A1"].value == "Metric"
    workbook.close()


def test_online_retail_demo_writes_report_workbook(tmp_path) -> None:
    input_path = tmp_path / "Online Retail.xlsx"
    output_path = tmp_path / "online_retail_business_report.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Online Retail"
    worksheet.append(["InvoiceNo", "StockCode", "Description", "Quantity", "InvoiceDate", "UnitPrice", "CustomerID", "Country"])
    worksheet.append(["536365", "85123A", "WHITE HANGING HEART T-LIGHT HOLDER", 6, datetime(2010, 12, 1, 8, 26), 2.55, 17850, "United Kingdom"])
    worksheet.append(["536365", "71053", "WHITE METAL LANTERN", 6, datetime(2010, 12, 1, 8, 26), 3.39, 17850, "United Kingdom"])
    worksheet.append(["C536383", "35004C", "SET OF 3 COLOURED FLYING DUCKS", -1, datetime(2010, 12, 1, 9, 49), 4.65, 15311, "United Kingdom"])
    workbook.save(input_path)

    build_online_retail_report(OnlineRetailReportConfig(input_path=input_path, output_path=output_path))

    report = load_workbook(output_path, read_only=True)
    assert "Executive Summary" in report.sheetnames
    assert "Monthly Revenue" in report.sheetnames
    assert "Cancelled Orders" in report.sheetnames
    assert "Data Quality Issues" in report.sheetnames

    summary_rows = list(report["Executive Summary"].iter_rows(values_only=True))
    report.close()
    summary = dict(summary_rows[1:])
    assert summary["Raw rows processed"] == 3
    assert summary["Valid sales rows"] == 2
    assert summary["Cancelled / return rows"] == 1
