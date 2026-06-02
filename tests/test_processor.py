from __future__ import annotations

import json
from datetime import date

from file_processor.extractors import supported_extensions
from file_processor.processor import process_folder
from file_processor.tabular import write_tabular_outputs


def test_supported_extensions_include_core_formats() -> None:
    extensions = supported_extensions()
    assert ".txt" in extensions
    assert ".json" in extensions
    assert ".csv" in extensions
    assert ".xlsx" in extensions
    assert ".docx" in extensions
    assert ".pptx" in extensions


def test_process_folder_writes_json_results(tmp_path) -> None:
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "note.txt").write_text("hello ETL", encoding="utf-8")
    (input_dir / "data.json").write_text('{"name": "sample"}', encoding="utf-8")

    results = process_folder(input_dir, output_dir)

    assert len(results) == 2
    assert all(result.status == "ok" for result in results)
    output_files = sorted(output_dir.glob("*.json"))
    assert len(output_files) == 2

    payload = json.loads(output_files[0].read_text(encoding="utf-8"))
    assert payload["source_path"]
    assert payload["metadata"]["size_bytes"] > 0


def test_process_folder_serializes_excel_dates(tmp_path) -> None:
    from openpyxl import Workbook

    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ClientInfo"
    worksheet.append(["Field", "Value"])
    worksheet.append(["ReviewDate", date(2026, 5, 15)])
    workbook.save(input_dir / "client.xlsx")

    results = process_folder(input_dir, output_dir)

    assert len(results) == 1
    assert results[0].status == "ok"
    payload = json.loads(next(output_dir.glob("*.json")).read_text(encoding="utf-8"))
    assert payload["content"]["sheets"]["ClientInfo"][1][1] == "2026-05-15T00:00:00"


def test_tabular_xlsx_output_combines_workbook_sheets(tmp_path) -> None:
    from openpyxl import Workbook, load_workbook

    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()

    workbook = Workbook()
    client_info = workbook.active
    client_info.title = "ClientInfo"
    client_info.append(["Field", "Value"])
    client_info.append(["ClientID", "CL-1001"])
    client_info.append(["ClientName", "Alpha Growth Trust"])
    holdings = workbook.create_sheet("Holdings")
    holdings.append(["Ticker", "MarketValue"])
    holdings.append(["BHP.AX", 19236])
    workbook.save(input_dir / "client.xlsx")

    results = process_folder(input_dir, output_dir, write_json=False)
    output_paths = write_tabular_outputs(results, output_dir, "xlsx")

    assert output_paths == [output_dir / "processed_output.xlsx"]
    output_workbook = load_workbook(output_paths[0], read_only=True)
    assert "RunSummary" in output_workbook.sheetnames
    assert "Holdings" in output_workbook.sheetnames
    holdings_rows = list(output_workbook["Holdings"].iter_rows(values_only=True))
    assert "ClientID" in holdings_rows[0]
    assert holdings_rows[1][holdings_rows[0].index("ClientID")] == "CL-1001"


def test_tabular_csv_output_writes_one_file_per_table(tmp_path) -> None:
    import csv

    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    (input_dir / "transactions.csv").write_text("Ticker,Quantity\nVAS.AX,10\n", encoding="utf-8")

    results = process_folder(input_dir, output_dir, write_json=False)
    output_paths = write_tabular_outputs(results, output_dir, "csv")

    assert output_dir / "RunSummary.csv" in output_paths
    rows_path = output_dir / "Rows.csv"
    assert rows_path in output_paths
    with rows_path.open(newline="", encoding="utf-8") as file:
        rows = list(csv.DictReader(file))
    assert rows[0]["Ticker"] == "VAS.AX"
