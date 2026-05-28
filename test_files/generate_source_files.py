from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_DIR = Path(__file__).parent / "srcfiles"


FILES: list[dict[str, Any]] = [
    {
        "file": "client_alpha_growth_portfolio.xlsx",
        "info": [
            ["Field", "Value"],
            ["ClientID", "CL-1001"],
            ["ClientName", "Alpha Growth Trust"],
            ["Adviser", "M. Chen"],
            ["RiskProfile", "Growth"],
            ["BaseCurrency", "AUD"],
            ["ReviewDate", date(2026, 5, 15)],
        ],
        "holdings": [
            ["AccountNo", "AssetClass", "Ticker", "ISIN", "SecurityName", "Quantity", "Price", "MarketValue", "CostBase", "Currency"],
            ["A-7001", "Australian Equity", "BHP.AX", "AU000000BHP4", "BHP Group Ltd", 420, 45.80, 19236.00, 15700.00, "AUD"],
            ["A-7001", "Australian Equity", "CBA.AX", "AU000000CBA7", "Commonwealth Bank of Australia", 110, 121.40, 13354.00, 10250.00, "AUD"],
            ["A-7001", "ETF", "VAS.AX", "AU000000VAS1", "Vanguard Australian Shares ETF", 250, 99.70, 24925.00, 23100.00, "AUD"],
            ["A-7002", "International Equity", "AAPL", "US0378331005", "Apple Inc", 40, 190.20, 7608.00, 6850.00, "USD"],
        ],
        "transactions": [
            ["TradeDate", "AccountNo", "Type", "Ticker", "Quantity", "Price", "GrossAmount", "Fees", "NetAmount", "Currency"],
            [date(2026, 1, 12), "A-7001", "Buy", "VAS.AX", 50, 96.10, 4805.00, 14.95, 4819.95, "AUD"],
            [date(2026, 2, 18), "A-7001", "Dividend", "BHP.AX", None, None, 462.00, 0.00, 462.00, "AUD"],
            [date(2026, 4, 3), "A-7002", "Buy", "AAPL", 10, 184.30, 1843.00, 9.99, 1852.99, "USD"],
        ],
        "cash": [["AccountNo", "Currency", "CashBalance"], ["A-7001", "AUD", 18542.55], ["A-7002", "USD", 2310.12]],
        "notes": [["Severity", "Issue", "Detail"], ["Info", "Clean file", "Typical growth client with AUD and USD assets"]],
    },
    {
        "file": "client_beta_balanced_portfolio.xlsx",
        "info": [
            ["Field", "Value"],
            ["ClientID", "CL-1002"],
            ["ClientName", "Beta Balanced Super"],
            ["Adviser", "S. Patel"],
            ["RiskProfile", "Balanced"],
            ["BaseCurrency", "AUD"],
            ["ReviewDate", date(2026, 5, 17)],
        ],
        "holdings": [
            ["AccountNo", "AssetClass", "Ticker", "ISIN", "SecurityName", "Quantity", "Price", "MarketValue", "CostBase", "Currency"],
            ["B-8101", "ETF", "VGS.AX", "AU000000VGS8", "Vanguard MSCI Index International Shares ETF", 300, 129.10, 38730.00, 35200.00, "AUD"],
            ["B-8101", "Fixed Income", "VAF.AX", "AU000000VAF4", "Vanguard Australian Fixed Interest ETF", 540, 45.30, 24462.00, 25100.00, "AUD"],
            ["B-8101", "Australian Equity", "TLS.AX", "AU000000TLS2", "Telstra Group Ltd", 2100, 3.92, 8232.00, 7700.00, "AUD"],
            ["B-8102", "Cash Fund", "AAA.AX", "AU00000AAA06", "Betashares Australian High Interest Cash ETF", 180, 50.08, 9014.40, 9000.00, "AUD"],
        ],
        "transactions": [
            ["TradeDate", "AccountNo", "Type", "Ticker", "Quantity", "Price", "GrossAmount", "Fees", "NetAmount", "Currency"],
            [date(2026, 1, 20), "B-8101", "Buy", "VAF.AX", 100, 45.80, 4580.00, 12.00, 4592.00, "AUD"],
            [date(2026, 3, 28), "B-8101", "Sell", "TLS.AX", 400, 4.05, 1620.00, 12.00, 1608.00, "AUD"],
            [date(2026, 4, 30), "B-8102", "Distribution", "AAA.AX", None, None, 91.20, 0.00, 91.20, "AUD"],
        ],
        "cash": [["AccountNo", "Currency", "CashBalance"], ["B-8101", "AUD", 6740.25], ["B-8102", "AUD", 1200.00]],
        "notes": [["Severity", "Issue", "Detail"], ["Info", "Balanced allocation", "Includes fixed income, equity, ETF, and cash fund positions"]],
    },
    {
        "file": "client_carrington_income_portfolio.xlsx",
        "info": [
            ["Field", "Value"],
            ["ClientID", "CL-1003"],
            ["ClientName", "Carrington Income Account"],
            ["Adviser", "L. Nguyen"],
            ["RiskProfile", "Income"],
            ["BaseCurrency", "AUD"],
            ["ReviewDate", date(2026, 5, 20)],
        ],
        "holdings": [
            ["AccountNo", "AssetClass", "Ticker", "ISIN", "SecurityName", "Quantity", "Price", "MarketValue", "CostBase", "Currency"],
            ["C-9201", "Fixed Income", "IAF.AX", "AU000000IAF9", "iShares Core Composite Bond ETF", 800, 104.20, 83360.00, 82000.00, "AUD"],
            ["C-9201", "Listed Property", "VAP.AX", "AU000000VAP1", "Vanguard Australian Property Securities Index ETF", 220, 84.70, 18634.00, 19600.00, "AUD"],
            ["C-9201", "Australian Equity", "WES.AX", "AU000000WES1", "Wesfarmers Ltd", 75, 67.50, 5062.50, 4200.00, "AUD"],
            ["C-9201", "Term Deposit", "TD-ANZ-0626", "", "ANZ Term Deposit Jun 2026", 1, 50000.00, 50000.00, 50000.00, "AUD"],
        ],
        "transactions": [
            ["TradeDate", "AccountNo", "Type", "Ticker", "Quantity", "Price", "GrossAmount", "Fees", "NetAmount", "Currency"],
            [date(2026, 2, 1), "C-9201", "Interest", "TD-ANZ-0626", None, None, 687.50, 0.00, 687.50, "AUD"],
            [date(2026, 2, 15), "C-9201", "Dividend", "WES.AX", None, None, 78.75, 0.00, 78.75, "AUD"],
            [date(2026, 5, 1), "C-9201", "Buy", "IAF.AX", 120, 103.80, 12456.00, 14.95, 12470.95, "AUD"],
        ],
        "cash": [["AccountNo", "Currency", "CashBalance"], ["C-9201", "AUD", 3844.80]],
        "notes": [["Severity", "Issue", "Detail"], ["Info", "Income-oriented client", "Contains a non-listed term deposit style holding"]],
    },
    {
        "file": "client_delta_edge_cases_portfolio.xlsx",
        "info": [
            ["Field", "Value"],
            ["ClientID", "CL-1004"],
            ["ClientName", "Delta Family Office"],
            ["Adviser", "R. Osei"],
            ["RiskProfile", "High Growth"],
            ["BaseCurrency", "AUD"],
            ["ReviewDate", date(2026, 5, 21)],
        ],
        "holdings": [
            ["AccountNo", "AssetClass", "Ticker", "ISIN", "SecurityName", "Quantity", "Price", "MarketValue", "CostBase", "Currency"],
            ["D-1101", "International Equity", "MSFT", "US5949181045", "Microsoft Corp", 55, 428.10, 23545.50, 18800.00, "USD"],
            ["D-1101", "International Equity", "", "US88160R1014", "Tesla Inc", 12, 177.80, 2133.60, 3900.00, "USD"],
            ["D-1101", "Australian Equity", "NAB.AX", "AU000000NAB4", "National Australia Bank Ltd", 0, 35.20, 0.00, 5200.00, "AUD"],
            ["D-1102", "Alternatives", "PE-FUND-07", "", "Private Equity Fund VII", 1, None, None, 75000.00, "AUD"],
        ],
        "transactions": [
            ["TradeDate", "AccountNo", "Type", "Ticker", "Quantity", "Price", "GrossAmount", "Fees", "NetAmount", "Currency"],
            ["not-a-date", "D-1101", "Buy", "MSFT", 15, 402.50, 6037.50, 19.95, 6057.45, "USD"],
            [date(2026, 3, 10), "D-1101", "Buy", "", 12, 177.80, 2133.60, 19.95, 2153.55, "USD"],
            [date(2026, 4, 22), "D-1102", "Capital Call", "PE-FUND-07", None, None, 25000.00, 0.00, 25000.00, "AUD"],
        ],
        "cash": [["AccountNo", "Currency", "CashBalance"], ["D-1101", "USD", -425.40], ["D-1102", "AUD", 0.00]],
        "notes": [
            ["Severity", "Issue", "Detail"],
            ["Warning", "Missing ticker", "Tesla row has ISIN but blank ticker"],
            ["Warning", "Bad date", "First transaction uses text instead of a date"],
            ["Warning", "Null value", "Private equity valuation price and market value are blank"],
        ],
    },
    {
        "file": "client_evergreen_multi_account_portfolio.xlsx",
        "info": [
            ["Field", "Value"],
            ["ClientID", "CL-1005"],
            ["ClientName", "Evergreen Charitable Foundation"],
            ["Adviser", "A. Rossi"],
            ["RiskProfile", "Conservative"],
            ["BaseCurrency", "AUD"],
            ["ReviewDate", date(2026, 5, 22)],
        ],
        "holdings": [
            ["AccountNo", "AssetClass", "Ticker", "ISIN", "SecurityName", "Quantity", "Price", "MarketValue", "CostBase", "Currency"],
            ["E-3301", "ETF", "VDHG.AX", "AU00000VDHG9", "Vanguard Diversified High Growth ETF", 150, 67.25, 10087.50, 9900.00, "AUD"],
            ["E-3301", "Fixed Income", "VBND.AX", "AU00000VBND0", "Vanguard Global Aggregate Bond Index ETF", 600, 47.15, 28290.00, 29100.00, "AUD"],
            ["E-3302", "Cash Fund", "BILL.AX", "AU00000BILL4", "iShares Core Cash ETF", 500, 100.21, 50105.00, 50000.00, "AUD"],
            ["E-3303", "International Equity", "GOOGL", "US02079K3059", "Alphabet Inc Class A", 25, 173.60, 4340.00, 4100.00, "USD"],
        ],
        "transactions": [
            ["TradeDate", "AccountNo", "Type", "Ticker", "Quantity", "Price", "GrossAmount", "Fees", "NetAmount", "Currency"],
            [date(2026, 1, 5), "E-3301", "Buy", "VBND.AX", 200, 48.20, 9640.00, 12.00, 9652.00, "AUD"],
            [date(2026, 3, 31), "E-3302", "Distribution", "BILL.AX", None, None, 310.44, 0.00, 310.44, "AUD"],
            [date(2026, 5, 5), "E-3303", "FX Transfer", "USD", None, None, 5000.00, 0.00, 5000.00, "USD"],
        ],
        "cash": [["AccountNo", "Currency", "CashBalance"], ["E-3301", "AUD", 980.44], ["E-3302", "AUD", 640.10], ["E-3303", "USD", 512.90]],
        "notes": [["Severity", "Issue", "Detail"], ["Info", "Multi-account file", "Conservative foundation with three account numbers and cross-currency cash"]],
    },
]


def add_sheet(workbook: Workbook, title: str, rows: list[list[Any]], table_name: str | None = None) -> None:
    worksheet = workbook.create_sheet(title)
    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(cell.value, float):
                cell.number_format = "#,##0.00"

    for column in worksheet.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 42)

    worksheet.freeze_panes = "A2"

    if table_name:
        ref = f"A1:{worksheet.cell(row=worksheet.max_row, column=worksheet.max_column).coordinate}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        worksheet.add_table(table)


def build_file(definition: dict[str, Any]) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "ClientInfo", definition["info"])
    add_sheet(workbook, "Holdings", definition["holdings"], "HoldingsTable")
    add_sheet(workbook, "Transactions", definition["transactions"], "TransactionsTable")
    add_sheet(workbook, "Cash", definition["cash"], "CashTable")
    add_sheet(workbook, "Notes", definition["notes"])

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / definition["file"]
    workbook.save(output_path)
    return output_path


def main() -> None:
    for definition in FILES:
        print(build_file(definition))


if __name__ == "__main__":
    main()
